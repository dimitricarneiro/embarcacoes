# Importações do Flask
from flask import Blueprint, request, jsonify, render_template, Response, send_file, redirect, url_for, make_response, flash
from flask import current_app as app
from app import limiter

# Flask-Login (Autenticação)
from flask_login import login_required, current_user

# Banco de Dados e Modelos
from app import db
from app.models import PedidoAutorizacao, Usuario, Notificacao, Embarcacao, Veiculo, Pessoa, Equipamento, Exigencia, Alerta, Prorrogacao

# Segurança
from app.security import role_required

# Formulários
from app.forms import AlertaForm, PedidoSearchForm

# Utilitários
import io
from datetime import datetime, date, timedelta
import csv
import re
import logging

# SQLAlchemy
from sqlalchemy.sql import func
from sqlalchemy.orm import joinedload

# Bibliotecas para Gerar Relatórios PDF
from io import BytesIO
from reportlab.lib.pagesizes import A3, landscape, A4, portrait
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from app.utils import validar_cnpj, validar_cpf, normalizar_cnpj

# Bibliotecas para Gerar Planilhas Excel
from openpyxl import Workbook
from openpyxl.styles import Font

# Bibliotecas para Gerar QRcode
import uuid
import qrcode
import base64


################Funções auxiliares
def filtrar_pedidos(args, current_user):
    """
    Retorna um objeto Query já com todos os filtros e ordenação aplicada.
    
    Parâmetros:
    - args: dict (ex. request.args)
    - current_user: objeto User (com role e id)
    
    Filtros incluídos:
      • Distinção RFB × usuário comum
      • nome_empresa, cnpj_empresa, status
      • data_inicio, data_termino, data_criacao
      • nome_embarcacao (via join com Embarcacao)
      • ordenação decrescente por ID
    """
    form = PedidoSearchForm(args)

    # 1) Base da query
    if current_user.role == "RFB":
        query = PedidoAutorizacao.query
    else:
        query = PedidoAutorizacao.query.filter_by(usuario_id=current_user.id)

    # 2) Aplicar filtros do form
    if form.nome_empresa.data:
        query = query.filter(
            PedidoAutorizacao.empresa_responsavel.ilike(f"%{form.nome_empresa.data}%")
        )

    if form.cnpj_empresa.data:
        query = query.filter(
            PedidoAutorizacao.cnpj_empresa.ilike(f"%{form.cnpj_empresa.data}%")
        )

    if form.status.data:
        query = query.filter(PedidoAutorizacao.status == form.status.data)

    if form.data_inicio.data:
        query = query.filter(PedidoAutorizacao.data_inicio >= form.data_inicio.data)

    if form.data_termino.data:
        query = query.filter(PedidoAutorizacao.data_termino <= form.data_termino.data)

    if form.data_criacao.data:
        query = query.filter(
            func.date(PedidoAutorizacao.data_criacao_pedido) == form.data_criacao.data
        )

    if form.nome_embarcacao.data:
        query = (
            query
            .join(PedidoAutorizacao.embarcacoes)
            .filter(
                func.lower(Embarcacao.nome) ==
                form.nome_embarcacao.data.lower()
            )
        )

    # 3) Ordenação final
    return query.order_by(PedidoAutorizacao.id.desc())

def verificar_alertas(novo_pedido):
    """
    Verifica se o novo pedido atende a algum alerta cadastrado e, se sim, cria notificações para os respectivos usuários RFB.
    """
    alertas = Alerta.query.filter_by(ativo=True).all()
    for alerta in alertas:
        # Se o alerta for do tipo "embarcacao", verificamos cada embarcação do pedido
        if alerta.tipo == "embarcacao":
            for embarcacao in novo_pedido.embarcacoes:
                if alerta.valor.lower() in embarcacao.nome.lower():
                    mensagem = (f"Novo pedido {novo_pedido.id} contém embarcação com nome "
                                f"correspondente ao seu alerta ('{alerta.valor}').")
                    criar_notificacao(alerta.usuario_id, mensagem)
                    break  # Evita notificar o mesmo alerta mais de uma vez para o mesmo pedido

        # Se o alerta for do tipo "cnpj", comparamos com o CNPJ da empresa do pedido
        elif alerta.tipo == "cnpj":
            if normalizar_cnpj(novo_pedido.cnpj_empresa) == normalizar_cnpj(alerta.valor):
                mensagem = (
                    f"Novo pedido {novo_pedido.id} criado pelo CNPJ "
                    f"'{novo_pedido.cnpj_empresa}' corresponde ao seu alerta."
                )
                criar_notificacao(alerta.usuario_id, mensagem)

        # Se o alerta for do tipo "cpf", percorremos cada pessoa associada ao pedido
        elif alerta.tipo == "cpf":
            for pessoa in novo_pedido.pessoas:
                cpf_ped = re.sub(r"\D", "", pessoa.cpf or "")
                if cpf_ped == alerta.valor:
                    mensagem = (
                        f"Novo pedido {novo_pedido.id} contém pessoa com CPF "
                        f"'{pessoa.cpf}'."
                    )
                    criar_notificacao(alerta.usuario_id, mensagem)
                    break  # evita múltiplas notificações do mesmo alerta

        # Se o alerta for do tipo "meio_de_transporte", checamos substring no campo
        elif alerta.tipo == "meio_de_transporte":
            meio = (novo_pedido.meio_de_transporte or "").lower()
            if alerta.valor.lower() in meio:
                mensagem = (
                    f"Novo pedido {novo_pedido.id} com meio de transporte "
                    f"'{novo_pedido.meio_de_transporte}' corresponde ao seu alerta."
                )
                criar_notificacao(alerta.usuario_id, mensagem)

def criar_notificacao(usuario_id, mensagem):
    """ Cria uma notificação para um usuário específico """
    nova_notificacao = Notificacao(usuario_id=usuario_id, mensagem=mensagem)
    db.session.add(nova_notificacao)
    db.session.commit()

def gerar_token():
    # Pode ser um UUID ou um hash baseado em dados e uma secret key
    return str(uuid.uuid4())

def gerar_qr_code(token):
    # Gera a URL para verificação usando url_for (assumindo que você tem a rota 'verificar_comprovante')
    url = url_for('pedidos.verificar_comprovante', token=token, _external=True)
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    # Converter a imagem para base64 para exibir no template
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")
    return img_str

def validar_horario(horario):
    """
    Verifica se o horário está no formato HH:MM e se os valores são válidos.

    Args:
        horario (str): Horário no formato "HH:MM".

    Returns:
        bool: True se o horário for válido, False caso contrário.
    """
    if not isinstance(horario, str):
        return False

    # Expressão regular para validar formato HH:MM
    padrao_horario = re.compile(r"^(0[0-9]|1[0-9]|2[0-3]):([0-5][0-9])$")

    return bool(padrao_horario.match(horario))

def validar_placa(placa):
    """
    Valida a placa do veículo considerando dois padrões:
    - Padrão antigo: 3 letras seguidas de 4 números (ex.: ABC1234)
    - Padrão Mercosul: 3 letras, 1 número, 1 letra e 2 números (ex.: ABC1D23)

    Args:
        placa (str): Placa do veículo.

    Returns:
        bool: True se a placa for válida, False caso contrário.
    """
    if not isinstance(placa, str):
        return False

    # Remove hífens e converte para maiúsculo
    placa = placa.replace("-", "").upper()

    padrao_antigo = re.compile(r'^[A-Z]{3}[0-9]{4}$')
    padrao_mercosul = re.compile(r'^[A-Z]{3}[0-9][A-Z][0-9]{2}$')

    return bool(padrao_antigo.match(placa)) or bool(padrao_mercosul.match(placa))

######### Início de definição das rotas ##########################################################################

pedidos_bp = Blueprint('pedidos', __name__)

@pedidos_bp.route('/')
def home():
    """Redireciona usuários não logados para a página de login"""
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))  # Redireciona para login
    return redirect(url_for('pedidos.exibir_pedidos'))  # Se logado, vai para /lista-pedidos

@pedidos_bp.route('/api/pedidos-autorizacao', methods=['POST', 'GET'])
@login_required
@role_required("comum")
def gerenciar_pedidos():
    """
    POST: Cria um novo pedido de autorização de serviço.
    GET: Retorna todos os pedidos cadastrados com suporte a filtros, paginação e ordenação.
    """
   
    if request.method == 'POST':
        # Restrição: Somente usuários com role "comum" podem criar um novo pedido
        if current_user.role != "comum":
            return jsonify({"error": "Você não pode criar um novo pedido."}), 403

        try:
            # Obter dados em formato JSON
            data = request.get_json()
            if not data:
                return jsonify({"error": "Dados JSON não enviados ou inválidos."}), 400

            # Validação do CNPJ
            cnpj = data.get("cnpj_empresa", "")
            if not validar_cnpj(cnpj):
                return jsonify({"error": "CNPJ inválido!"}), 400

            # Validação do CNPJ da agência
            cnpj = data.get("cnpj_agencia", "")
            if not validar_cnpj(cnpj):
                return jsonify({"error": "CNPJ da agência inválido!"}), 400

            # Dicionário com os campos obrigatórios e seus rótulos amigáveis
            campos_obrigatorios = {
                "nome_empresa": "Nome da empresa",
                "cnpj_empresa": "CNPJ da empresa",
                "endereco_empresa": "Endereço da empresa",
                "motivo_solicitacao": "Motivo da solicitação",
                "data_inicio": "Data do início",
                "data_termino": "Data do término",
                "horario_inicio_servicos": "Horário de início dos serviços",
                "horario_termino_servicos": "Horário de término dos serviços",
                "certificado_livre_pratica": "Certificado livre prática",
                "cidade_servico": "Cidade do serviço",
                "embarcacoes": "Embarcações",
                "pessoas": "Pessoas",
                "termo_responsabilidade": "Termo de responsabilidade"
            }

            # Verificação dos campos que não foram preenchidos
            campos_invalidos = [
                label 
                for campo, label in campos_obrigatorios.items() 
                if not data.get(campo)
            ]

            if campos_invalidos:
                return jsonify({
                    "error": f"Campos obrigatórios: {', '.join(campos_invalidos)}"
                }), 400

            # Verifica se o termo de responsabilidade foi aceito
            if not data.get("termo_responsabilidade", False):
                return jsonify({"error": "Você precisa aceitar os termos de responsabilidade."}), 400

            # Conversão das datas para objetos date
            try:
                data_inicio = datetime.strptime(data["data_inicio"], "%Y-%m-%d").date()
                data_termino = datetime.strptime(data["data_termino"], "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "Formato de data inválido. Use 'YYYY-MM-DD'."}), 400

            hoje = datetime.today().date()
            if data_inicio < hoje:
                return jsonify({"error": "A data de início deve ser hoje ou uma data futura."}), 400
            
            # Verifica se a data de início está dentro dos próximos 90 dias
            if data_inicio > hoje + timedelta(days=90):
                return jsonify({"error": "A data de início deve estar dentro dos próximos 90 dias."}), 400
            
            if data_termino < data_inicio:
                return jsonify({"error": "A data de término deve ser maior ou igual à data de início."}), 400

            # Validação da duração máxima do serviço: máximo 5 dias
            duracao = (data_termino - data_inicio).days
            if duracao > 5:
                return jsonify({"error": "A duração máxima do serviço é de 5 dias."}), 400

            # Validação de horário de início e término dos serviços
            horario_inicio = data.get("horario_inicio_servicos", "").strip()
            horario_termino = data.get("horario_termino_servicos", "").strip()

            if not validar_horario(horario_inicio):
                return jsonify({"error": "Horário de início inválido! Use o formato HH:MM e valores entre 00:00 e 23:59."}), 400

            if not validar_horario(horario_termino):
                return jsonify({"error": "Horário de término inválido! Use o formato HH:MM e valores entre 00:00 e 23:59."}), 400
                
            # Verifica tamanho máximo dos campos
            empresa_responsavel=data["nome_empresa"]
            if len(empresa_responsavel) > 255:
                return jsonify({"error": "O nome da empresa deve ter até 255 caraceres."}), 400
            
            certificado_livre_pratica = data["certificado_livre_pratica"]
            if len(certificado_livre_pratica) > 16:
                return jsonify({"error": "O certificado de livre prática deve ter até 16 caraceres."}), 400            

            # Criação do novo pedido, incluindo os novos campos
            novo_pedido = PedidoAutorizacao(
                empresa_responsavel=empresa_responsavel,
                cnpj_empresa=data["cnpj_empresa"],
                endereco_empresa=data["endereco_empresa"],
                motivo_solicitacao=data["motivo_solicitacao"],
                data_inicio=data_inicio,
                data_termino=data_termino,
                horario_inicio_servicos=data["horario_inicio_servicos"],
                horario_termino_servicos=data["horario_termino_servicos"],
                certificado_livre_pratica=certificado_livre_pratica,
                cidade_servico=data["cidade_servico"],
                observacoes=data.get("observacoes", None),
                agencia_maritima=data.get("agencia_maritima", None),
                cnpj_agencia=data.get("cnpj_agencia", None),
                representante_agencia=data.get("representante_agencia", None),
                meio_de_transporte=data.get("meio_de_transporte", None),
                termo_responsabilidade=data.get("termo_responsabilidade", False),
                usuario_id=current_user.id
            )

            # -------------------------------
            # Processamento de Embarcações
            # -------------------------------
            for embarcacao_data in data["embarcacoes"]:
                nome = embarcacao_data.get("nome", "").strip()
                if nome:
                    embarcacao = db.session.query(Embarcacao).filter_by(nome=nome).first()
                    if not embarcacao:
                        embarcacao = Embarcacao(
                            nome=nome,
                            imo=embarcacao_data.get("imo", "").strip(),
                            bandeira=embarcacao_data.get("bandeira", "").strip()
                        )
                        db.session.add(embarcacao)
                    else:
                        if embarcacao_data.get("imo"):
                            embarcacao.imo = embarcacao_data.get("imo").strip()
                        if embarcacao_data.get("bandeira"):
                            embarcacao.bandeira = embarcacao_data.get("bandeira").strip()
                    novo_pedido.embarcacoes.append(embarcacao)

            # -------------------------------
            # Processamento de Equipamentos
            # -------------------------------
            if "equipamentos" in data and data["equipamentos"]:
                for equipamento_data in data["equipamentos"]:
                    descricao = equipamento_data.get("descricao", "").strip()
                    numero_serie = equipamento_data.get("numero_serie", "").strip()
                    quantidade = equipamento_data.get("quantidade", 0)
                    unidade = equipamento_data.get("unidade", "").strip()
                    if descricao and numero_serie and quantidade:
                        equipamento = db.session.query(Equipamento).filter_by(numero_serie=numero_serie).first()
                        if not equipamento:
                            equipamento = Equipamento(
                                descricao=descricao,
                                numero_serie=numero_serie,
                                quantidade=int(quantidade),
                                unidade=unidade
                            )
                            db.session.add(equipamento)
                        else:
                            equipamento.quantidade = int(quantidade)
                            equipamento.unidade = unidade
                        novo_pedido.equipamentos.append(equipamento)

            # -------------------------------
            # Processamento de Pessoas
            # -------------------------------
            
            # Verifica duplicidade de CPFs nas pessoas
            cpfs = [
                pessoa_data.get("cpf", "").strip()
                for pessoa_data in data.get("pessoas", [])
                if pessoa_data.get("cpf", "").strip()
            ]
            if len(cpfs) != len(set(cpfs)):
                return jsonify({"error": "Há CPFs duplicados na lista de pessoas."}), 400
            
            for pessoa_data in data["pessoas"]:
                nome_pessoa = pessoa_data.get("nome", "").strip()
                cpf_pessoa = pessoa_data.get("cpf", "").strip()
                
                # Valida o CPF de cada pessoa
                if not validar_cpf(cpf_pessoa):
                    return jsonify({"error": f"CPF {cpf_pessoa} inválido!"}), 400

                isps = pessoa_data.get("isps", "").strip()
                funcao = pessoa_data.get("funcao", "").strip()
                local_embarque = pessoa_data.get("local_embarque", "").strip()
                local_desembarque = pessoa_data.get("local_desembarque", "").strip()

                if nome_pessoa and cpf_pessoa:
                    pessoa = db.session.query(Pessoa).filter_by(cpf=cpf_pessoa).first()
                    if not pessoa:
                        pessoa = Pessoa(
                            nome=nome_pessoa,
                            cpf=cpf_pessoa,
                            isps=isps,
                            funcao=funcao,
                            local_embarque=local_embarque,
                            local_desembarque=local_desembarque
                        )
                        db.session.add(pessoa)
                    else:
                        if isps:
                            pessoa.isps = isps
                        if funcao:
                            pessoa.funcao = funcao
                        if local_embarque:
                            pessoa.local_embarque = local_embarque
                        if local_desembarque:
                            pessoa.local_desembarque = local_desembarque
                    novo_pedido.pessoas.append(pessoa)

            # -------------------------------
            # Processamento de Veículos
            # -------------------------------
            if "veiculos" in data and data["veiculos"]:
                for veiculo_data in data["veiculos"]:
                    modelo_veiculo = veiculo_data.get("modelo", "").strip()
                    placa_veiculo = veiculo_data.get("placa", "").strip()
                    if modelo_veiculo and placa_veiculo:
                        # Valida a placa antes de continuar
                        if not validar_placa(placa_veiculo):
                            return jsonify({"error": f"Placa do veículo '{placa_veiculo}' é inválida. Use o formato AAA1234 ou AAA1A23."}), 400

                        veiculo = db.session.query(Veiculo).filter_by(placa=placa_veiculo).first()
                        if not veiculo:
                            veiculo = Veiculo(modelo=modelo_veiculo, placa=placa_veiculo)
                            db.session.add(veiculo)
                        novo_pedido.veiculos.append(veiculo)

            db.session.add(novo_pedido)
            db.session.commit()

            # Notificações
            # Este trecho de código está comentado, pois a chamada para criar uma notificação foi movida para a rota da agência
            #verificar_alertas(novo_pedido)
            #administradores = Usuario.query.filter_by(role="RFB").all()
            #mensagem = f"Novo pedido {novo_pedido.id} foi cadastrado e aguarda aprovação."
            #for admin in administradores:
            #    criar_notificacao(admin.id, mensagem)

            return jsonify({
                "id": novo_pedido.id,  # Inclui o id do pedido criado
                "redirect_url": url_for('pedidos.exibir_pedidos')
            }), 200

        except Exception as e:
            logging.exception("Erro ao criar pedido de autorização:")
            return jsonify({"error": "Ocorreu um erro interno no servidor."}), 500

    #TODO: avaliar a necessidade desse método GET.
    #Acredito que ele não esteja mais sendo usado e possa ser apagado.
    elif request.method == 'GET':
        try:
            query = PedidoAutorizacao.query
            page = request.args.get("page", default=1, type=int)
            per_page = request.args.get("per_page", default=10, type=int)
            query = query.order_by(PedidoAutorizacao.data_inicio.desc())
            pedidos_paginados = query.paginate(page=page, per_page=per_page, error_out=False)

            pedidos_lista = []
            for pedido in pedidos_paginados.items:
                pedido_dict = {
                    "id_autorizacao": pedido.id,
                    "nome_empresa": pedido.empresa_responsavel,
                    "cnpj_empresa": pedido.cnpj_empresa,
                    "endereco_empresa": pedido.endereco_empresa,
                    "motivo_solicitacao": pedido.motivo_solicitacao,
                    "data_inicio_servico": pedido.data_inicio.strftime("%Y-%m-%d"),
                    "data_termino_servico": pedido.data_termino.strftime("%Y-%m-%d"),
                    "horario_servicos": f"{pedido.horario_inicio_servicos} - {pedido.horario_termino_servicos}"
                }
                pedidos_lista.append(pedido_dict)

            return jsonify({
                "total_pedidos": pedidos_paginados.total,
                "pagina_atual": pedidos_paginados.page,
                "total_paginas": pedidos_paginados.pages,
                "pedidos": pedidos_lista
            }), 200

        except Exception as e:
            logging.exception("Erro ao recuperar pedidos de autorização:")
            return jsonify({"error": "Ocorreu um erro interno no servidor."}), 500

@pedidos_bp.route('/pedido/<int:pedido_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required("comum")
def editar_pedido(pedido_id):
    """
    Rota para editar um pedido existente via formulário web.
    Somente o criador pode editar e apenas se o status for 'pendente'.
    """
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Verifica que o pedido que está sendo editado foi criado pelo usuário logado
    if pedido.usuario_id != current_user.id:
        flash("Você não tem permissão para editar esse pedido.", "danger")
        return redirect(url_for('pedidos.exibir_pedidos'))

    if pedido.status != "aguardando_agencia":
        flash("Somente pedidos que ainda não foram confirmados pela agência podem ser editados.", "warning")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    if request.method == 'POST':
        # Obter os dados do formulário (incluindo os novos campos do pedido)
        nome_empresa = request.form.get('nome_empresa')
        cnpj_empresa = request.form.get('cnpj_empresa')
        endereco_empresa = request.form.get('endereco_empresa')
        motivo_solicitacao = request.form.get('motivo_solicitacao')
        data_inicio_str = request.form.get('data_inicio')
        data_termino_str = request.form.get('data_termino')
        horario_inicio = request.form.get('horario_inicio_servicos')
        horario_termino = request.form.get('horario_termino_servicos')
        certificado = request.form.get('certificado_livre_pratica')
        cidade_servico = request.form.get('cidade_servico')
        observacoes = request.form.get('observacoes')
        agencia_maritima = request.form.get('agencia_maritima')
        cnpj_agencia = request.form.get('cnpj_agencia')
        representante_agencia = request.form.get('representante_agencia')
        meio_de_transporte = request.form.get('meio_de_transporte')
        termo_responsabilidade = request.form.get('termo_responsabilidade')
        # Converte o valor do checkbox para booleano
        termo_responsabilidade = True if termo_responsabilidade in ["on", "true", "True", "1"] else False

        try:
            data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
            data_termino = datetime.strptime(data_termino_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Formato de data inválido. Use 'YYYY-MM-DD'.", "danger")
            return render_template('formulario.html', pedido=pedido)

        # Atualiza os atributos do pedido
        pedido.empresa_responsavel = nome_empresa
        pedido.cnpj_empresa = cnpj_empresa
        pedido.endereco_empresa = endereco_empresa
        pedido.motivo_solicitacao = motivo_solicitacao
        pedido.data_inicio = data_inicio
        pedido.data_termino = data_termino
        pedido.horario_inicio_servicos = horario_inicio
        pedido.horario_termino_servicos = horario_termino
        pedido.certificado_livre_pratica = certificado
        pedido.cidade_servico = cidade_servico
        pedido.observacoes = observacoes
        pedido.agencia_maritima = agencia_maritima
        pedido.cnpj_agencia = cnpj_agencia
        pedido.representante_agencia = representante_agencia
        pedido.meio_de_transporte = meio_de_transporte
        pedido.termo_responsabilidade = termo_responsabilidade

        db.session.commit()
        flash("Pedido atualizado com sucesso!", "success")

        return jsonify({
            "redirect_url": url_for('pedidos.exibir_pedidos')
        }), 200

    return render_template('formulario.html', pedido=pedido)

@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>', methods=['PUT'])
@login_required
@role_required("comum")
def atualizar_pedido_api(pedido_id):
    """
    Atualiza um pedido via API.
    Somente o criador e pedidos com status 'pendente' podem ser editados.
    """
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    if pedido.usuario_id != current_user.id:
        return jsonify({"error": "Você não tem permissão para editar esse pedido."}), 403

    if pedido.status != "aguardando_agencia":
        return jsonify({"error": "Somente pedidos que ainda não foram confirmados pela agência podem ser editados."}), 400

    data = request.get_json()
    if not data:
        return jsonify({"error": "Nenhum dado enviado."}), 400

    try:
        # Atualiza os campos principais
        pedido.empresa_responsavel = data.get("nome_empresa")
        pedido.cnpj_empresa = data.get("cnpj_empresa")
        pedido.endereco_empresa = data.get("endereco_empresa")
        pedido.motivo_solicitacao = data.get("motivo_solicitacao")
        pedido.data_inicio = datetime.strptime(data.get("data_inicio"), "%Y-%m-%d").date()
        pedido.data_termino = datetime.strptime(data.get("data_termino"), "%Y-%m-%d").date()

        # Validação da duração máxima do serviço: máximo 5 dias
        duracao = (pedido.data_termino - pedido.data_inicio).days
        if duracao > 5:
            return jsonify({"error": "A duração máxima do serviço é de 5 dias."}), 400

        pedido.horario_inicio_servicos = data.get("horario_inicio_servicos")
        pedido.horario_termino_servicos = data.get("horario_termino_servicos")
        pedido.certificado_livre_pratica = data.get("certificado_livre_pratica")
        pedido.cidade_servico = data.get("cidade_servico")
        pedido.observacoes = data.get("observacoes")
        pedido.agencia_maritima = data.get("agencia_maritima")
        pedido.cnpj_agencia = data.get("cnpj_agencia")
        pedido.representante_agencia = data.get("representante_agencia")
        pedido.meio_de_transporte = data.get("meio_de_transporte")
        pedido.termo_responsabilidade = data.get("termo_responsabilidade", False)

        # -------------------------------
        # Atualização de Embarcações
        # -------------------------------
        pedido.embarcacoes.clear()
        for embarcacao_data in data.get("embarcacoes", []):
            nome = embarcacao_data.get("nome", "").strip()
            if nome:
                embarcacao = db.session.query(Embarcacao).filter_by(nome=nome).first()
                if not embarcacao:
                    embarcacao = Embarcacao(
                        nome=nome,
                        imo=embarcacao_data.get("imo", "").strip(),
                        bandeira=embarcacao_data.get("bandeira", "").strip()
                    )
                    db.session.add(embarcacao)
                else:
                    if embarcacao_data.get("imo"):
                        embarcacao.imo = embarcacao_data.get("imo").strip()
                    if embarcacao_data.get("bandeira"):
                        embarcacao.bandeira = embarcacao_data.get("bandeira").strip()
                pedido.embarcacoes.append(embarcacao)

        # -------------------------------
        # Atualização de Veículos
        # -------------------------------
        pedido.veiculos.clear()
        for veiculo_data in data.get("veiculos", []):
            modelo = veiculo_data.get("modelo", "").strip()
            placa = veiculo_data.get("placa", "").strip()
            if modelo and placa:
                veiculo = db.session.query(Veiculo).filter_by(placa=placa).first()
                if not veiculo:
                    veiculo = Veiculo(modelo=modelo, placa=placa)
                    db.session.add(veiculo)
                pedido.veiculos.append(veiculo)

        # -------------------------------
        # Atualização de Equipamentos
        # -------------------------------
        pedido.equipamentos.clear()
        for equipamento_data in data.get("equipamentos", []):
            descricao = equipamento_data.get("descricao", "").strip()
            numero_serie = equipamento_data.get("numero_serie", "").strip()
            quantidade = equipamento_data.get("quantidade", 0)
            unidade = equipamento_data.get("unidade", "").strip()
            if descricao and numero_serie and quantidade:
                equipamento = db.session.query(Equipamento).filter_by(numero_serie=numero_serie).first()
                if not equipamento:
                    equipamento = Equipamento(
                        descricao=descricao,
                        numero_serie=numero_serie,
                        quantidade=int(quantidade),
                        unidade=unidade
                    )
                    db.session.add(equipamento)
                else:
                    equipamento.quantidade = int(quantidade)
                    equipamento.unidade=unidade
                pedido.equipamentos.append(equipamento)

        # -------------------------------
        # Atualização de Pessoas
        # -------------------------------
        pedido.pessoas.clear()
        for pessoa_data in data.get("pessoas", []):
            nome = pessoa_data.get("nome", "").strip()
            cpf = pessoa_data.get("cpf", "").strip()
            isps = pessoa_data.get("isps", "").strip()
            funcao = pessoa_data.get("funcao", "").strip()
            local_embarque = pessoa_data.get("local_embarque", "").strip()
            local_desembarque = pessoa_data.get("local_desembarque", "").strip()
            if nome and cpf:
                pessoa = db.session.query(Pessoa).filter_by(cpf=cpf).first()
                if not pessoa:
                    pessoa = Pessoa(
                        nome=nome,
                        cpf=cpf,
                        isps=isps,
                        funcao=funcao,
                        local_embarque=local_embarque,
                        local_desembarque=local_desembarque
                    )
                    db.session.add(pessoa)
                else:
                    if isps:
                        pessoa.isps = isps
                    if funcao:
                        pessoa.funcao = funcao
                    if local_embarque:
                        pessoa.local_embarque = local_embarque
                    if local_desembarque:
                        pessoa.local_desembarque = local_desembarque
                pedido.pessoas.append(pessoa)

        db.session.commit()
        return jsonify({
            "redirect_url": url_for('pedidos.exibir_pedidos')
        }), 200

    except Exception as e:
        db.session.rollback()
        logging.exception("Erro ao atualizar pedido:")
        return jsonify({"error": "Erro ao atualizar pedido."}), 500

@pedidos_bp.route('/pedido/<int:pedido_id>/prorrogar', methods=['GET', 'POST'])
@login_required
@role_required("comum")
def prorrogar_pedido(pedido_id):
    """
    Rota para solicitar a prorrogação de um pedido.
    
    Regras:
    - Somente o criador do pedido pode solicitar a prorrogação.
    - O pedido deve ter status "aprovado".
    - A solicitação só pode ser feita se faltarem menos de 3 dias para o término.
    - Não é permitida a solicitação se já houver uma prorrogação pendente.
    - Por padrão, o status da prorrogação é 'pendente'.
    """
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Verifica se o usuário logado é o criador do pedido
    if pedido.usuario_id != current_user.id:
        flash("Você não tem permissão para prorrogar este pedido.", "danger")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Verifica se o pedido está aprovado
    if pedido.status != "aprovado":
        flash("Somente pedidos aprovados podem ser prorrogados.", "warning")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Verifica se faltam menos de 3 dias para a data de término
    dias_restantes = (pedido.data_termino - date.today()).days
    if dias_restantes >= 3:
        flash("A prorrogação só pode ser solicitada quando faltam menos de 3 dias para o término.", "warning")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Verifica se já existe uma prorrogação pendente para esse pedido
    if any(prorrogacao.status_prorrogacao == 'pendente' for prorrogacao in pedido.prorrogacoes):
        flash("Você já solicitou uma prorrogação que ainda está pendente.", "warning")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Se for GET, exibe o formulário para solicitar a prorrogação
    if request.method == 'GET':
        return render_template('formulario_prorrogacao.html', pedido=pedido)

    # Se for POST, processa os dados do formulário
    nova_data_str = request.form.get('data_termino_nova')
    justificativa = request.form.get('justificativa')

    if not nova_data_str:
        flash("A nova data de término é obrigatória.", "danger")
        return render_template('formulario_prorrogacao.html', pedido=pedido)
    
    if not justificativa:
        flash("A justificativa é obrigatória.", "danger")
        return render_template('formulario_prorrogacao.html', pedido=pedido)

    try:
        nova_data = datetime.strptime(nova_data_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Formato de data inválido. Use 'YYYY-MM-DD'.", "danger")
        return render_template('formulario_prorrogacao.html', pedido=pedido)

    # Opcional: a nova data deve ser posterior à data de término atual
    if nova_data <= pedido.data_termino:
        flash("A nova data deve ser posterior à data de término atual.", "danger")
        return render_template('formulario_prorrogacao.html', pedido=pedido)

    # Cria o registro da prorrogação com status padrão 'pendente'
    nova_prorrogacao = Prorrogacao(
        pedido_id=pedido.id,
        data_termino_antiga=pedido.data_termino,
        data_termino_nova=nova_data,
        status_prorrogacao='pendente',
        justificativa=justificativa
    )
    db.session.add(nova_prorrogacao)
    db.session.commit()

    flash("Pedido de prorrogação enviado com sucesso e está pendente de aprovação.", "success")
    return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>/prorrogacoes/<int:prorrogacao_id>/aprovar', methods=['PUT'])
@login_required
@role_required("RFB")
def aprovar_prorrogacao(pedido_id, prorrogacao_id):
    """
    Aprova a prorrogação de um pedido.
    
    Regras:
    - Apenas usuários com role "RFB" podem aprovar.
    - A prorrogação deve estar com status 'pendente'.
    - Ao aprovar, atualiza o campo 'data_termino' do pedido com a nova data.
    """
    # Verifica se o usuário tem permissão para aprovar
    if current_user.role != "RFB":
        return jsonify({"error": "Acesso não autorizado"}), 403

    # Busca o pedido de autorização
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Busca a prorrogação vinculada ao pedido
    prorrogacao = (
        Prorrogacao.query
        .filter_by(id=prorrogacao_id, pedido_id=pedido_id)
        .first_or_404()
    )

    # Verifica se a prorrogação já foi processada
    if prorrogacao.status_prorrogacao != "pendente":
        return jsonify({"error": "Esta prorrogação já foi processada."}), 400

    # Aprova a prorrogação: atualiza o status e a data de término do pedido
    prorrogacao.status_prorrogacao = "aprovada"
    pedido.data_termino = prorrogacao.data_termino_nova

    db.session.commit()

    return jsonify({
        "message": "Prorrogação aprovada com sucesso!",
        "id_prorrogacao": prorrogacao.id,
        "status_prorrogacao": prorrogacao.status_prorrogacao,
        "id_autorizacao": pedido.id,
        "data_termino_atualizada": pedido.data_termino.strftime("%Y-%m-%d")
    }), 200

@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>/prorrogacoes/<int:prorrogacao_id>/rejeitar', methods=['PUT'])
@login_required
@role_required("RFB")
def rejeitar_prorrogacao(pedido_id, prorrogacao_id):
    """
    Rota para rejeitar uma prorrogação de pedido.
    
    Regras:
    - Apenas usuários com role "RFB" podem rejeitar.
    - A prorrogação deve estar com status 'pendente'.
    """
    # Verifica se o usuário logado tem permissão para rejeitar (apenas role "RFB")
    if current_user.role != "RFB":
        return jsonify({"error": "Acesso não autorizado"}), 403

    # Busca o pedido de autorização
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Busca a prorrogação vinculada ao pedido
    prorrogacao = (
        Prorrogacao.query
        .filter_by(id=prorrogacao_id, pedido_id=pedido_id)
        .first_or_404()
    )

    # Verifica se a prorrogação já foi processada
    if prorrogacao.status_prorrogacao != "pendente":
        return jsonify({"error": "Esta prorrogação já foi processada."}), 400

    # Atualiza o status da prorrogação para 'rejeitada'
    prorrogacao.status_prorrogacao = "rejeitada"
    db.session.commit()

    return jsonify({
        "message": "Prorrogação rejeitada com sucesso!",
        "id_prorrogacao": prorrogacao.id,
        "status_prorrogacao": prorrogacao.status_prorrogacao,
        "id_autorizacao": pedido.id
    }), 200

@pedidos_bp.route('/lista-pedidos', methods=['GET'])
@login_required
@role_required("RFB", "comum")
def exibir_pedidos():
    """Exibe os pedidos em uma página HTML com filtros, busca e paginação."""
    
    # Monta o form (para manter os campos preenchidos no template)
    form = PedidoSearchForm(request.args)

    # Parâmetros de paginação
    page     = request.args.get("page",     default=1,  type=int)
    per_page = request.args.get("per_page", default=10, type=int)

    # Utiliza o helper para aplicar filtros + ordenação
    query = filtrar_pedidos(request.args, current_user)
    pedidos_paginados = query.paginate(page=page, per_page=per_page, error_out=False)

    hoje = date.today()
    
    return render_template('lista-pedidos.html', pedidos=pedidos_paginados, form=form, hoje=hoje)

@pedidos_bp.route('/pedido/<int:pedido_id>', methods=['GET'])
@login_required
def exibir_detalhes_pedido(pedido_id):
    """
    Exibe os detalhes de um pedido específico, incluindo:
      - exigências + quem as criou
      - quem analisou/aprovou o pedido (usuario_que_analisou)
    """
    pedido = (
        PedidoAutorizacao.query
        .options(
            joinedload(PedidoAutorizacao.exigencias)
                .joinedload(Exigencia.usuario),
            joinedload(PedidoAutorizacao.usuario_que_analisou)   # <— aqui
        )
        .get_or_404(pedido_id)
    )

    if current_user.role != 'RFB' and pedido.usuario_id != current_user.id:
        flash("Você não tem permissão para detalhar esse pedido.", "danger")
        return redirect(url_for('pedidos.exibir_pedidos'))

    return render_template('detalhes-pedido.html', pedido=pedido)
    
@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>/aprovar', methods=['PUT'])
@login_required
@role_required("RFB")
def aprovar_pedido(pedido_id):
    """ Aprova um pedido de autorização """

    # Verifica se o usuário tem permissão (apenas usuários com role "RFB" podem aprovar)
    if current_user.role != "RFB":
        return jsonify({"error": "Acesso não autorizado"}), 403

    # Busca o pedido no banco
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Verifica se o prazo final dos trabalho ainda não foi alcançado
    if pedido.data_termino < date.today():
        return jsonify({
            "error": "Pedido expirado."
        }), 400

    # Verifica se já foi aprovado
    if pedido.status == "aprovado":
        return jsonify({"error": "Este pedido já foi aprovado"}), 400

    # Aprova o pedido e registra as informações de análise
    pedido.status = "aprovado"
    pedido.data_analise_pedido = datetime.utcnow()
    pedido.id_usuario_que_analisou_pedido = current_user.id

    db.session.commit()

    return jsonify({
        "message": "Pedido aprovado com sucesso!",
        "id_autorizacao": pedido.id,
        "status": pedido.status
    }), 200

@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>/rejeitar', methods=['PUT'])
@login_required
@role_required("RFB")
def rejeitar_pedido(pedido_id):
    """ Rejeita um pedido de autorização """

    # Verifica se o usuário tem permissão para rejeitar (apenas usuários com role "RFB")
    if current_user.role != "RFB":
        return jsonify({"error": "Acesso não autorizado"}), 403

    # Busca o pedido no banco
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Verifica se o prazo final dos trabalho ainda não foi alcançado
    if pedido.data_termino < date.today():
        return jsonify({
            "error": "Pedido expirado."
        }), 400

    # Verifica se já foi aprovado ou rejeitado
    if pedido.status != "pendente":
        return jsonify({"error": f"Este pedido já foi {pedido.status}."}), 400

    # Rejeita o pedido e registra as informações de análise
    pedido.status = "rejeitado"
    pedido.data_analise_pedido = datetime.utcnow()
    pedido.id_usuario_que_analisou_pedido = current_user.id

    db.session.commit()

    return jsonify({
        "message": "Pedido rejeitado com sucesso!",
        "id_autorizacao": pedido.id,
        "status": pedido.status
    }), 200

@pedidos_bp.route('/api/pedidos-autorizacao/<int:pedido_id>/exigir', methods=['POST'])
@login_required
@role_required("RFB")
def exigir_pedido(pedido_id):
    """
    Registra uma exigência para um pedido de autorização.
    Apenas usuários com role "RFB" podem executar essa ação.
    Ao fazer a exigência, é necessário informar:
      - motivo_exigencia: o motivo da exigência;
      - prazo_exigencia: prazo (no formato AAAA-MM-DD) para o cumprimento da exigência.
    O status do pedido é atualizado para "exigência".
    """
    try:
        # Verifica se o usuário tem permissão
        if current_user.role != "RFB":
            return jsonify({"error": "Acesso não autorizado"}), 403

        # Busca o pedido no banco
        pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

        # Verifica se o prazo final dos trabalho ainda não foi alcançado
        if pedido.data_termino < date.today():
            return jsonify({
                "error": "Pedido expirado."
            }), 400

        # Verifica se o pedido está no status pendente
        if pedido.status != "pendente":
            return jsonify({"error": f"Este pedido já foi {pedido.status}."}), 400

        # Obtém os dados enviados no corpo da requisição (JSON)
        data = request.get_json()
        if not data:
            return jsonify({"error": "Dados inválidos"}), 400

        motivo_exigencia = data.get('motivo_exigencia')
        prazo_exigencia = data.get('prazo_exigencia')

        # Valida se os campos obrigatórios foram informados
        if not motivo_exigencia or not prazo_exigencia:
            return jsonify({"error": "Os campos 'motivo_exigencia' e 'prazo_exigencia' são obrigatórios."}), 400

        # Converte o prazo para o formato de data (AAAA-MM-DD)
        try:
            prazo_exigencia_date = datetime.strptime(prazo_exigencia, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Formato de prazo_exigencia inválido. Utilize AAAA-MM-DD."}), 400

        # Verifica se o prazo_exigencia não é anterior à data atual
        if prazo_exigencia_date < date.today():
            return jsonify({"error": "O prazo não pode ser anterior a hoje."}), 400

        # Cria o registro da exigência, salvando também o id do usuário que fez a exigência
        exigencia = Exigencia(
            pedido_id=pedido.id,
            motivo_exigencia=motivo_exigencia,
            prazo_exigencia=prazo_exigencia_date,
            usuario_id=current_user.id  # Salva o id do usuário que fez a exigência
        )
        db.session.add(exigencia)

        # Atualiza o status do pedido para "exigência"
        pedido.status = "exigência"
        db.session.commit()

        return jsonify({
            "message": "Exigência registrada com sucesso!",
            "id_autorizacao": pedido.id,
            "status": pedido.status
        }), 200

    except Exception as e:
        app.logger.error("Erro ao registrar exigência", exc_info=e)
        return jsonify({"error": "Erro interno do servidor"}), 500

@pedidos_bp.route('/exigencia/<int:exigencia_id>')
@login_required
@role_required("comum", "RFB")
def detalhes_exigencia(exigencia_id):
    exigencia = Exigencia.query.get_or_404(exigencia_id)

    # Verifica se o usuário atual é 'RFB' ou se é o criador do pedido associado à exigência
    if current_user.role != 'RFB' and exigencia.pedido.usuario_id != current_user.id:
        return jsonify({"error": "Você não tem permissão para ver essa exigência."}), 403

    # Passa também o pedido associado à exigência para o template
    return render_template('detalhes_exigencia.html', exigencia=exigencia, pedido=exigencia.pedido)

@pedidos_bp.route('/pedido/<int:pedido_id>/responder-exigencia/<int:exigencia_id>', methods=['GET', 'POST'])
@login_required
@role_required("comum")
def responder_exigencia(pedido_id, exigencia_id):
    """
    Rota para que o criador do pedido responda uma exigência.
    
    Regras:
    - Apenas o usuário que criou o pedido pode responder a exigência.
    - A resposta deve ser enviada até (ou na) data do prazo_exigencia.
    - Não é permitida a resposta se já existir um texto_resposta cadastrado.
    - Ao responder, o pedido.status é atualizado para "pendente".
    """
    # Busca o pedido e a exigência
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)
    exigencia = Exigencia.query.filter_by(id=exigencia_id, pedido_id=pedido.id).first_or_404()

    # Verifica se o usuário logado é o criador do pedido
    if pedido.usuario_id != current_user.id:
        flash("Você não tem permissão para responder esta exigência.", "danger")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Se a exigência já foi respondida, não permite nova resposta
    if exigencia.texto_resposta:
        flash("Esta exigência já foi respondida.", "warning")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Verifica se o prazo para resposta ainda não expirou
    if date.today() > exigencia.prazo_exigencia:
        flash("O prazo para responder esta exigência já expirou.", "danger")
        return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

    # Se for GET, exibe o template com os detalhes e o formulário de resposta
    if request.method == 'GET':
        return render_template('detalhes_exigencia.html', pedido=pedido, exigencia=exigencia)

    # Se for POST, processa o formulário enviado
    texto_resposta = request.form.get('texto_resposta')
    if not texto_resposta:
        flash("O campo de resposta é obrigatório.", "danger")
        return render_template('detalhes_exigencia.html', pedido=pedido, exigencia=exigencia)

    # Registra a resposta com data atual e atualiza o status do pedido
    exigencia.texto_resposta = texto_resposta
    exigencia.data_resposta = datetime.utcnow()
    pedido.status = "pendente"

    db.session.commit()

    flash("Exigência respondida com sucesso e o pedido foi atualizado para 'pendente'.", "success")
    return redirect(url_for('pedidos.exibir_detalhes_pedido', pedido_id=pedido.id))

@pedidos_bp.route('/formulario-pedido', methods=['GET'])
@login_required  # Apenas usuários logados podem acessar
def exibir_formulario():
    """ Rota que exibe o formulário para preencher o pedido de autorização """
    return render_template('formulario.html')

@pedidos_bp.route('/admin')
@login_required
@role_required("RFB")
def admin_dashboard():
    """ Painel Administrativo - Somente para usuários RFB """

    # Verifica se o usuário é RFB
    if current_user.role != "RFB":
        return redirect(url_for("pedidos.exibir_pedidos"))

    # Estatísticas gerais
    total_pedidos = PedidoAutorizacao.query.count()
    pedidos_aprovados = PedidoAutorizacao.query.filter_by(status="aprovado").count()
    pedidos_rejeitados = PedidoAutorizacao.query.filter_by(status="rejeitado").count()
    pedidos_pendentes = PedidoAutorizacao.query.filter_by(status="pendente").count()
    total_usuarios = Usuario.query.count()

    # Contagem de pedidos por dia
    pedidos_por_dia = (
        db.session.query(func.date(PedidoAutorizacao.data_inicio), func.count())
        .group_by(func.date(PedidoAutorizacao.data_inicio))
        .order_by(func.date(PedidoAutorizacao.data_inicio))
        .all()
    )

    # Preparar dados para o gráfico
    datas = [str(p[0]) for p in pedidos_por_dia]
    pedidos_quantidade = [p[1] for p in pedidos_por_dia]

    return render_template("admin.html", 
                           total_pedidos=total_pedidos, 
                           pedidos_aprovados=pedidos_aprovados,
                           pedidos_rejeitados=pedidos_rejeitados, 
                           pedidos_pendentes=pedidos_pendentes,
                           total_usuarios=total_usuarios,
                           datas=datas, 
                           pedidos_quantidade=pedidos_quantidade)

@pedidos_bp.route('/admin/alertas', methods=['GET', 'POST'])
@login_required
@role_required("RFB")
def gerenciar_alertas():
    """Exibe os alertas do usuário RFB e permite a criação de novos alertas."""
    if current_user.role != "RFB":
        return redirect(url_for('pedidos.exibir_pedidos'))
    
    form = AlertaForm()
    if form.validate_on_submit():
        # Normaliza o valor se for alerta de CPF (remove pontos e traços)
        valor = form.valor.data
        if form.tipo.data == "cpf":
            valor = re.sub(r"\D", "", valor)

        novo_alerta = Alerta(
            usuario_id=current_user.id,
            tipo=form.tipo.data,
            valor=valor
        )
        db.session.add(novo_alerta)
        db.session.commit()
        return redirect(url_for("pedidos.gerenciar_alertas"))
    
    alertas = Alerta.query.filter_by(usuario_id=current_user.id).all()
    return render_template("gerenciar_alertas.html", alertas=alertas, form=form)

@pedidos_bp.route('/comprovante/<int:pedido_id>')
@login_required
def gerar_comprovante(pedido_id):
    pedido = PedidoAutorizacao.query.get_or_404(pedido_id)

    # Verifica se o pedido foi aprovado e se o usuário tem permissão para visualizar o comprovante
    if pedido.status != "aprovado" or pedido.usuario_id != current_user.id:
        return jsonify({"error": "Você não tem permissão para visualizar este comprovante."}), 403

    # Se ainda não existir um token, gere e salve
    if not pedido.token_comprovante:
        pedido.token_comprovante = gerar_token()
        db.session.commit()

    qr_code_base64 = gerar_qr_code(pedido.token_comprovante)

    return render_template('comprovante.html', pedido=pedido, qr_code=qr_code_base64)

@pedidos_bp.route('/verificar-comprovante/<string:token>')
def verificar_comprovante(token):
    # Procura um pedido com o token informado
    pedido = PedidoAutorizacao.query.filter_by(token_comprovante=token, status="aprovado").first()
    if not pedido:
        return render_template('comprovante_invalido.html'), 404

    # Se válido, exiba os dados do pedido
    return render_template('detalhes_comprovante.html', pedido=pedido)

@pedidos_bp.route('/admin/exportar-csv')
@login_required
@role_required("RFB", "comum")
def exportar_csv():
    """Exporta os pedidos como um arquivo CSV."""

    # 1) Busca filtrada de pedidos via helper (aplica filtros e ordenação)
    query = filtrar_pedidos(request.args, current_user)
    pedidos = query.all()

    # 2) Prepara buffer e BOM
    output = io.StringIO()
    output.write('\ufeff')  # BOM para Excel

    # 3) Cria o writer
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    # 4) Cabeçalho
    writer.writerow([
        "ID", "Empresa", "CNPJ", "Motivo",
        "Data Início", "Data Término",
        "Data Solicitação", "Data Análise",
        "RFB Servidor Analisou", "Status"
    ])

    # 5) Linhas de dados
    for p in pedidos:
        writer.writerow([
            p.id,
            p.empresa_responsavel,
            p.cnpj_empresa,
            p.motivo_solicitacao,
            p.data_inicio.strftime("%d/%m/%Y"),
            p.data_termino.strftime("%d/%m/%Y"),
            p.data_criacao_pedido.strftime("%d/%m/%Y"),
            p.data_analise_pedido.strftime("%d/%m/%Y") if p.data_analise_pedido else "",
            p.usuario_que_analisou.username if p.usuario_que_analisou else "",
            p.status
        ])

    # 6) Finaliza o buffer
    csv_content = output.getvalue()
    output.close()

    # 7) Monta a resposta
    response = make_response(csv_content)
    response.headers["Content-Disposition"] = (
        "attachment; filename=relatorio_pedidos.csv"
    )
    response.headers["Content-Type"] = "text/csv; charset=utf-8"

    return response

@pedidos_bp.route('/admin/exportar-pdf')
@login_required
@role_required("RFB", "comum")
def exportar_pdf():
    """Exporta os pedidos como PDF, incluindo sumário estatístico."""

    # 1) Busca filtrada de pedidos via helper (aplica filtros e ordenação)
    query = filtrar_pedidos(request.args, current_user)
    pedidos = query.all()

    # 2) Estatísticas
    total_pedidos     = len(pedidos)
    pedidos_aprovados = sum(1 for p in pedidos if p.status == "aprovado")
    pedidos_rejeitados= sum(1 for p in pedidos if p.status == "rejeitado")
    pedidos_pendentes = sum(1 for p in pedidos if p.status == "pendente")

    # Função utilitária para formatar datas seguras
    def fmt_date(dt):
        return dt.strftime("%d/%m/%Y") if dt else ""

    # 3) Montagem do PDF
    buffer = BytesIO()
    pdf    = SimpleDocTemplate(buffer, pagesize=landscape(A3))
    elementos = []

    styles = getSampleStyleSheet()
    elementos.append(Paragraph("<b>Relatório de Pedidos</b>", styles['Title']))
    elementos.append(Spacer(1, 12))

    # Cabeçalho da tabela
    dados = [[
        "ID", "Empresa", "CNPJ", "Motivo",
        "Data Início", "Data Término",
        "Data Solicitação", "Data Análise RFB",
        "Servidor Analisou", "Status"
    ]]
    for p in pedidos:
        dados.append([
            p.id,
            p.empresa_responsavel,
            p.cnpj_empresa,
            p.motivo_solicitacao,
            fmt_date(p.data_inicio),
            fmt_date(p.data_termino),
            fmt_date(p.data_criacao_pedido),
            fmt_date(p.data_analise_pedido),
            p.usuario_que_analisou.username if p.usuario_que_analisou else "",
            p.status
        ])

    tabela = Table(dados)
    tabela.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),    colors.grey),
        ('TEXTCOLOR',     (0,0),  (-1,0),    colors.whitesmoke),
        ('FONTNAME',      (0,0),  (-1,0),    'Helvetica-Bold'),
        ('ALIGN',         (0,0),  (-1,-1),   'CENTER'),

        ('FONTSIZE',      (0,0),  (-1,-1),   7),

        ('LEFTPADDING',   (0,0),  (-1,-1),   1),
        ('RIGHTPADDING',  (0,0),  (-1,-1),   1),
        ('TOPPADDING',    (0,0),  (-1,-1),   1),
        ('BOTTOMPADDING', (0,0),  (-1,-1),   1),

        ('BACKGROUND',    (0,1),  (-1,-1),   colors.beige),
        ('GRID',          (0,0),  (-1,-1),   0.5, colors.black),
    ]))
    elementos.append(tabela)
    elementos.append(Spacer(1, 20))

    # Sumário estatístico
    elementos.append(Paragraph("<b>Sumário Estatístico</b>", styles['Heading2']))
    elementos.append(Paragraph(f"Total de Pedidos: {total_pedidos}", styles['Normal']))
    elementos.append(Paragraph(f"Pedidos Aprovados: {pedidos_aprovados}", styles['Normal']))
    elementos.append(Paragraph(f"Pedidos Rejeitados: {pedidos_rejeitados}", styles['Normal']))
    elementos.append(Paragraph(f"Pedidos Pendentes: {pedidos_pendentes}", styles['Normal']))

    pdf.build(elementos)
    buffer.seek(0)

    # 4) Envia o arquivo — ajuste attachment_filename / download_name
    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_pedidos.pdf",
        mimetype="application/pdf"
    )

# nova rota para “PDF Completo”
from flask import request, send_file
from flask_login import login_required, current_user
from app.security import role_required
from app.forms import PedidoSearchForm
from app.models import PedidoAutorizacao
from io import BytesIO
import re

@pedidos_bp.route('/admin/exportar-pdf-completo')
@login_required
@role_required("RFB", "comum")
def exportar_pdf_completo():
    """Exporta um PDF completo, com cards detalhados de cada pedido."""
    # 1) Busca os pedidos filtrados
    query = filtrar_pedidos(request.args, current_user)
    pedidos = query.all()

    # 2) Prepara o documento
    buffer = BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=40, rightMargin=40,
        topMargin=40, bottomMargin=40
    )
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    label_style = ParagraphStyle(
        name='Label',
        parent=styles['Normal'],
        fontName='Helvetica-Bold'
    )

    elementos = []
    elementos.append(Paragraph("Relatório Completo de Pedidos", styles['Title']))
    elementos.append(Spacer(1, 12))

    # Helper para formatar CPF
    def fmt_cpf(cpf):
        digits = re.sub(r"\D", "", cpf or "")
        if len(digits) == 11:
            return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
        return cpf or ""

    # 3) Monta um “card” (tabela) para cada pedido
    for p in pedidos:
        data = []
        # — Campos básicos —
        data.extend([
            [Paragraph("ID do pedido:", label_style), Paragraph(str(p.id), normal)],
            [Paragraph("Empresa:", label_style), Paragraph(p.empresa_responsavel, normal)],
            [Paragraph("CNPJ:", label_style), Paragraph(p.cnpj_empresa, normal)],
            [Paragraph("Motivo:", label_style), Paragraph(p.motivo_solicitacao, normal)],
            [Paragraph("Data de início:", label_style),
             Paragraph(p.data_inicio.strftime("%d-%m-%Y"), normal)],
            [Paragraph("Data de término:", label_style),
             Paragraph(p.data_termino.strftime("%d-%m-%Y"), normal)],
            [Paragraph("Horário de início:", label_style),
             Paragraph(p.horario_inicio_servicos or "N/A", normal)],
            [Paragraph("Horário de término:", label_style),
             Paragraph(p.horario_termino_servicos or "N/A", normal)],
            [Paragraph("Agência Marítima:", label_style),
             Paragraph(p.agencia_maritima or "N/A", normal)],
            [Paragraph("CNPJ da Agência:", label_style),
             Paragraph(p.cnpj_agencia or "N/A", normal)],
            [Paragraph("Representante da Agência:", label_style),
             Paragraph(p.representante_agencia or "N/A", normal)],
            [Paragraph("Meio de Transporte:", label_style),
             Paragraph(p.meio_de_transporte or "N/A", normal)],
            [Paragraph("Situação:", label_style),
             Paragraph({
                 "pendente": "Pendente de análise pela RFB ⏳",
                 "aguardando_agencia": "Aguardando agência ⏳",
                 "aprovado": f"Aprovado ✅{f' (por {p.usuario_que_analisou.username})' if p.usuario_que_analisou else ''}",
                 "rejeitado": "Rejeitado ❌",
                 "exigência": "Com exigência ⚠",
                 "rejeitado_agencia": "Rejeitado pela agência ❌"
             }.get(p.status, p.status), normal)],
        ])

        # — Exigências —
        if p.exigencias:
            data.append([Paragraph("Exigências:", label_style), Paragraph("", normal)])
            for ex in p.exigencias:
                data.extend([
                    [Paragraph("Motivo da exigência:", label_style),
                     Paragraph(ex.motivo_exigencia, normal)],
                    [Paragraph("Prazo da exigência:", label_style),
                     Paragraph(ex.prazo_exigencia.strftime("%d/%m/%Y"), normal)],
                    [Paragraph("Data de criação:", label_style),
                     Paragraph(f"{ex.data_criacao.strftime('%d/%m/%Y %H:%M')} (por {ex.usuario.username})", normal)],
                    [Paragraph("Resposta em:", label_style),
                     Paragraph(ex.data_resposta.strftime("%d/%m/%Y %H:%M") if ex.data_resposta else "–", normal)],
                    [Paragraph("Texto da resposta:", label_style),
                     Paragraph(ex.texto_resposta or "Nenhuma resposta", normal)],
                ])
        else:
            data.append([Paragraph("Exigências:", label_style),
                         Paragraph("Nenhuma exigência registrada.", normal)])

        # — Embarcações —
        if p.embarcacoes:
            data.append([Paragraph("Embarcações:", label_style), Paragraph("", normal)])
            for em in p.embarcacoes:
                data.extend([
                    [Paragraph("Nome:", label_style), Paragraph(em.nome, normal)],
                    [Paragraph("IMO:", label_style), Paragraph(em.imo or "N/A", normal)],
                    [Paragraph("Bandeira:", label_style), Paragraph(em.bandeira or "N/A", normal)],
                ])
        else:
            data.append([Paragraph("Embarcações:", label_style),
                         Paragraph("Nenhuma embarcação associada.", normal)])

        # — Equipamentos —
        if p.equipamentos:
            data.append([Paragraph("Equipamentos:", label_style), Paragraph("", normal)])
            for eq in p.equipamentos:
                data.extend([
                    [Paragraph("Descrição:", label_style), Paragraph(eq.descricao, normal)],
                    [Paragraph("Número de Série:", label_style), Paragraph(eq.numero_serie, normal)],
                    [Paragraph("Quantidade:", label_style), Paragraph(str(eq.quantidade), normal)],
                    [Paragraph("Unidade:", label_style), Paragraph(eq.unidade or "N/A", normal)],
                ])
        else:
            data.append([Paragraph("Equipamentos:", label_style),
                         Paragraph("Nenhum equipamento associado.", normal)])

        # — Pessoas —
        if p.pessoas:
            data.append([Paragraph("Pessoas:", label_style), Paragraph("", normal)])
            for ps in p.pessoas:
                data.extend([
                    [Paragraph("Nome:", label_style), Paragraph(ps.nome, normal)],
                    [Paragraph("CPF:", label_style), Paragraph(fmt_cpf(ps.cpf), normal)],
                    [Paragraph("ISPS:", label_style), Paragraph(ps.isps or "–", normal)],
                    [Paragraph("Função:", label_style), Paragraph(ps.funcao or "–", normal)],
                    [Paragraph("Local de Embarque:", label_style), Paragraph(ps.local_embarque or "–", normal)],
                    [Paragraph("Local de Desembarque:", label_style), Paragraph(ps.local_desembarque or "–", normal)],
                ])
        else:
            data.append([Paragraph("Pessoas:", label_style),
                         Paragraph("Nenhuma pessoa associada.", normal)])

        # — Veículos —
        if p.veiculos:
            data.append([Paragraph("Veículos:", label_style), Paragraph("", normal)])
            for v in p.veiculos:
                data.extend([
                    [Paragraph("Modelo:", label_style), Paragraph(v.modelo, normal)],
                    [Paragraph("Placa:", label_style), Paragraph(v.placa, normal)],
                ])
        else:
            data.append([Paragraph("Veículos:", label_style),
                         Paragraph("Nenhum veículo associado.", normal)])

        # Monta a tabela “card”
        tabela = Table(data, colWidths=[140, 340], hAlign='LEFT')
        tabela.setStyle(TableStyle([
            ('BOX',          (0,0),  (-1,-1),   1,   colors.black),
            ('INNERGRID',    (0,0),  (-1,-1),   0.5, colors.grey),
            ('VALIGN',       (0,0),  (-1,-1),   'TOP'),
            ('BACKGROUND',   (0,0),  (-1,0),    colors.lightgrey),
            ('LEFTPADDING',  (0,0),  (-1,-1),   6),
            ('RIGHTPADDING', (0,0),  (-1,-1),   6),
            ('TOPPADDING',   (0,0),  (-1,-1),   4),
            ('BOTTOMPADDING',(0,0),  (-1,-1),   4),
        ]))
        elementos.append(tabela)
        elementos.append(Spacer(1, 16))

    # 4) Gera e retorna o PDF
    doc.build(elementos)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_pedidos_completo.pdf",
        mimetype="application/pdf"
    )

@pedidos_bp.route('/admin/exportar-excel')
@login_required
@role_required("RFB", "comum")
def exportar_excel():
    """ 
    Exporta os pedidos cadastrados como um arquivo Excel (.xlsx).
    
    O arquivo conterá duas planilhas:
      1. 'Pedidos': Relação completa dos pedidos.
      2. 'Sumário': Estatísticas gerais dos pedidos.
    """
    # Apenas usuários com role "RFB" podem acessar esta rota.
    #if current_user.role != "RFB":
    #    return redirect(url_for("pedidos.exibir_pedidos"))

    #pedidos = PedidoAutorizacao.query.all() #caso queira exportar todos os pedidos sem nenhum filtro
    
    # Busca filtrada de pedidos via helper (aplica filtros e ordenação)
    query = filtrar_pedidos(request.args, current_user)
    pedidos = query.all()

    # Calcular estatísticas
    total_pedidos = len(pedidos)
    pedidos_aprovados = len([p for p in pedidos if p.status == "aprovado"])
    pedidos_rejeitados = len([p for p in pedidos if p.status == "rejeitado"])
    pedidos_pendentes = len([p for p in pedidos if p.status == "pendente"])

    # Criar um novo Workbook
    wb = Workbook()

    # --------------------------
    # Planilha de Pedidos
    # --------------------------
    ws_pedidos = wb.active
    ws_pedidos.title = "Pedidos"

    # Cabeçalho da tabela
    headers = ["ID", "Empresa", "CNPJ", "Motivo", "Data Início", "Data Término", "Data Solicitação", "Data Análise", "RFB Servidor Analisou" , "Status"]
    ws_pedidos.append(headers)

    # Estilizar o cabeçalho (fonte em negrito)
    for col in range(1, len(headers) + 1):
        ws_pedidos.cell(row=1, column=col).font = Font(bold=True)

    # Preencher os dados de cada pedido
    for pedido in pedidos:
        # Formatar datas como "dd/mm/aaaa"
        data_inicio = pedido.data_inicio.strftime("%d/%m/%Y") if pedido.data_inicio else ""
        data_termino = pedido.data_termino.strftime("%d/%m/%Y") if pedido.data_termino else ""
        data_analise = pedido.data_analise_pedido.strftime("%d/%m/%Y") if pedido.data_analise_pedido else ""
        ws_pedidos.append([
            pedido.id,
            pedido.empresa_responsavel,
            pedido.cnpj_empresa,
            pedido.motivo_solicitacao,
            data_inicio,
            data_termino,
            pedido.data_criacao_pedido.strftime("%d/%m/%Y"),
            data_analise,
            pedido.usuario_que_analisou.username if pedido.usuario_que_analisou else "",
            pedido.status
        ])

    # Ajustar a largura das colunas para melhor visualização
    for column_cells in ws_pedidos.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Obter a letra da coluna
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws_pedidos.column_dimensions[column].width = adjusted_width

    # --------------------------
    # Planilha do Sumário Estatístico
    # --------------------------
    ws_sumario = wb.create_sheet(title="Sumário")
    
    # Título do sumário
    ws_sumario["A1"] = "Sumário Estatístico"
    ws_sumario["A1"].font = Font(bold=True, size=14)
    
    # Preencher estatísticas
    ws_sumario["A3"] = "Total de Pedidos:"
    ws_sumario["B3"] = total_pedidos

    ws_sumario["A4"] = "Pedidos Aprovados:"
    ws_sumario["B4"] = pedidos_aprovados

    ws_sumario["A5"] = "Pedidos Rejeitados:"
    ws_sumario["B5"] = pedidos_rejeitados

    ws_sumario["A6"] = "Pedidos Pendentes:"
    ws_sumario["B6"] = pedidos_pendentes

    # Ajustar a largura das colunas do sumário
    ws_sumario.column_dimensions["A"].width = 25
    ws_sumario.column_dimensions["B"].width = 15

    # --------------------------
    # Salvar o Workbook em um objeto BytesIO
    # --------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Retornar o arquivo para download
    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio_pedidos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@pedidos_bp.route("/api/notificacoes", methods=["GET"])
@limiter.limit("30 per minute")  # Permite 30 requisições por minuto só para essa rota
@login_required
@role_required("RFB")
def listar_notificacoes():
    """ Retorna notificações não lidas do usuário autenticado """
    notificacoes = Notificacao.query.filter_by(usuario_id=current_user.id, lida=False).all()

    return jsonify([
        {"id": n.id, "mensagem": n.mensagem, "data": n.data_criacao.strftime("%d/%m/%Y %H:%M"), "lida": n.lida}
        for n in notificacoes
    ])

@pedidos_bp.route('/api/notificacoes/<int:notificacao_id>/marcar-lida', methods=['PUT'])
@login_required
@role_required("RFB")
def marcar_notificacao_lida(notificacao_id):
    """ Marca uma notificação como lida """
    
    notificacao = Notificacao.query.filter_by(id=notificacao_id, usuario_id=current_user.id).first()

    if not notificacao:
        return jsonify({"error": "Notificação não encontrada"}), 404

    notificacao.lida = True
    db.session.commit()

    return jsonify({"message": "Notificação marcada como lida"}), 200
